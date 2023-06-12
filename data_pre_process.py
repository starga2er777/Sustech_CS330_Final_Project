import os
import openpyxl
from openpyxl.styles import PatternFill

# 打开工作簿（XLSX文件）
workbook = openpyxl.load_workbook('Copy of sections.xlsx')

# 选择工作表
worksheet = workbook['Sheet1'] 

with open('output.txt', 'w') as file:
    # 遍历行
    for row in worksheet.iter_rows():
        # 遍历单元格
        for cell in row:
            # 写入单元格的值到文本文件
            file.write(str(cell.value) + ' ') 
        file.write('\n')  # 写入换行符表示行结束

# 关闭工作簿
workbook.close()